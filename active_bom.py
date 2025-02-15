#!/usr/bin/env uv run

import argparse
import csv
import json
import re
import time
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tabulate import tabulate

CLIENT_ID = "fUmWDtTt1M7ZCDr6mrFK2upi4Cbe04LD"
CLIENT_SECRET = "IL4HaP2js9m5ircd"

RESISTOR_VALUE_MAP_ERJ = {
    "1.5kΩ": "1501",
    "10kΩ": "1002",
    "1kΩ": "1001",
    "26.1Ω": "26R1",
    "4.7kΩ": "4701",
    "220Ω": "2200",
    "510Ω": "5100",
    "5.6kΩ": "5601",
    "100kΩ": "1003",
    "60.4Ω": "60R4",
    "1.2kΩ": "1201",
    "180Ω": "1800",
}

RESISTOR_FOOTPRINT_MAP_ERJ = {
    "0402": "ERJ-2RKF",
    "0603": "ERJ-3EKF",
}

RESISTOR_FOOTPRINT_MAP_ERJ_SUFFIX = {
    "0402": "X",
    "0603": "V",
}

CAPACITOR_MAP = {
    (16, "100nF", "X7R", 10, "0402"): "GCM155R71C104KA55J",
    (50, "2.2uF", "X5R", 10, "0805"): "GCM21BR71C225KA64L",
    (6.3, "2.2uF", "X5R", 20, "0402"): "GRM155R61C225KE11D",
    (50, "12nF", "X7R", 10, "0402"): "GCM155R71E123KA55J",
    (50, "18pF", "C0G", 5, "0402"): "GCM1555C1H180JA16D",
    (50, "20pF", "C0G", 5, "0402"): "GCM1555C1H220JA16J",
    (25, "1uF", "X5R", 10, "0402"): "GRM155R61E105KA12D",
    (50, "4.7nF", "X7R", 10, "0402"): "GCM155R71H472KA37J",
}

MPN_MAP = {
    "C2040": "RP2040",
    "RP2040": "SC0914(13)",
    "C9002": "X322512MSB4SI",
    "C97521": "W25Q128JVSIQ",
    "X322512MSB4SI": "ECS-2333-120-BN-TR",
    "ERM8-040-05.0-X-DV-L-K-TR": "ERM8-040-05.0-L-DV-L-K-TR",
}


def get_digikey_token(client_id, client_secret):
    """Get DigiKey API token, using cached token if still valid."""
    cache_file = Path.home() / "Library" / "Caches" / "digikey" / "token.json"
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    if cache_file.exists():
        with open(cache_file) as f:
            token_data = json.load(f)
            if time.time() < token_data["expires_at"] - 60:
                return token_data["access_token"]

    # Get new token
    response = requests.post(
        "https://api.digikey.com/v1/oauth2/token",
        data={
            "client_id": client_id,
            "client_secret": client_secret,
            "grant_type": "client_credentials",
        },
    )
    response.raise_for_status()
    token_data = response.json()
    token_data["expires_at"] = time.time() + token_data["expires_in"]
    with open(cache_file, "w") as f:
        json.dump(token_data, f)
    return token_data["access_token"]


def get_cached_digikey_response(url, headers, json_data):
    cache_dir = Path.home() / "Library" / "Caches" / "digikey" / "search"
    mpn = json_data["Keywords"]  # Extract from search params
    cache_file = cache_dir / f"{normalize_filename(mpn)}.json"
    cache_dir.mkdir(parents=True, exist_ok=True)

    # Check cache
    if cache_file.exists():
        with open(cache_file) as f:
            cached_data = json.load(f)
            if time.time() < cached_data["cached_at"] + 3600 * 4:  # 4 hour cache
                return cached_data["response"]

    print(f"Fetching {mpn} from DigiKey API...")
    response = requests.post(url, headers=headers, json=json_data)
    if response.status_code == 200:
        cache_data = {
            "response": response.json(),
            "cached_at": time.time(),
            "request": {"url": url, "method": "POST", "json_data": json_data},
        }
        with open(cache_file, "w") as f:
            json.dump(cache_data, f, indent=2)
        return response.json()
    else:
        raise Exception(f"API request failed: {response.text}")


def normalize_filename(mpn):
    """Convert MPN to a valid filename by replacing invalid characters."""
    # Replace invalid filename characters with underscores
    valid_name = re.sub(r'[/<>:"\\|?*,]', "_", mpn)
    # Remove any multiple underscores
    valid_name = re.sub(r"_+", "_", valid_name)
    # Remove leading/trailing underscores
    valid_name = valid_name.strip("_")
    return valid_name


def extract_product_info(response, order_quantity):
    # print(json.dumps(response, indent=4))
    mpn = response["ManufacturerProductNumber"]
    description = response["Description"]["ProductDescription"]
    manufacturer = response["Manufacturer"]["Name"]
    data = {
        "mpn": mpn,
        "description": description,
        "manufacturer": manufacturer,
        "vendor": "DigiKey",
    }
    for param in response["Parameters"]:
        if param["ParameterText"] == "Supplier Device Package":
            data["footprint"] = param["ValueText"]
            break

    # find the product variant with the best unit price and moq < order_quantity
    best_unit_price = float("inf")
    for variation in response["ProductVariations"]:
        digikey_pn = variation["DigiKeyProductNumber"]
        available = int(variation["QuantityAvailableforPackageType"])
        moq = int(variation["MinimumOrderQuantity"])
        if available == 0 or moq > order_quantity:
            continue
        for pricing in variation["StandardPricing"]:
            break_quantity = int(pricing["BreakQuantity"])
            if break_quantity > order_quantity:
                continue
            unit_price = float(pricing["UnitPrice"])
            if unit_price < best_unit_price:
                best_unit_price = unit_price
                data["vendor_part_number"] = digikey_pn
                data["unit_price"] = best_unit_price
                data["available"] = available

    return data


def search_digikey_info(mpn, order_quantity):
    access_token = get_digikey_token(CLIENT_ID, CLIENT_SECRET)
    url = "https://api.digikey.com/products/v4/search/keyword"
    headers = {
        "accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-DIGIKEY-Client-Id": CLIENT_ID,
        "Content-Type": "application/json",
    }
    response = get_cached_digikey_response(
        url,
        headers,
        json_data={"Keywords": mpn, "Limit": 2},
    )
    if response["ExactMatches"]:
        product = response["ExactMatches"][0]
    elif response["Products"]:
        if len(response["Products"]) > 1:
            raise Exception(f"Multiple products found for MPN: {mpn}")
        product = response["Products"][0]
    else:
        raise Exception(f"No products found for MPN: {mpn}")
    return extract_product_info(product, order_quantity)


def erj_mpn(value, footprint):
    erj_value = RESISTOR_VALUE_MAP_ERJ[value]
    erj_prefix = RESISTOR_FOOTPRINT_MAP_ERJ[footprint]
    erj_suffix = RESISTOR_FOOTPRINT_MAP_ERJ_SUFFIX[footprint]
    return f"{erj_prefix}{erj_value}{erj_suffix}"


def parse_resistor_comment(comment):
    value_match = re.search(r"±1% ([0-9.]+[kΩ]+)", comment)
    footprint_match = re.search(r"([0-9]{4})", comment)
    if value_match and footprint_match:
        value = value_match.group(1)
        footprint = footprint_match.group(1)
        mpn = erj_mpn(value, footprint)
        comment = f"{value} ±1% {footprint} thick film resistor 0.1W 50V ±100ppm/C"
        data = {
            "mpn": mpn,
            "value": value,
            "footprint": footprint,
            "description": comment,
            "vendor": "DigiKey",
            "manufacturer": "Panasonic Electronic Components",
        }
        return data
    raise ValueError("Invalid resistor comment")


def parse_capacitor_comment(comment):
    voltage_match = re.search(r"(\d+\.?\d*)V", comment)
    value_match = re.search(r"(\d+\.?\d*)(pF|nF|uF)", comment)
    type_match = re.search(r"(X7R|X5R|C0G)", comment)
    tolerance_match = re.search(r"±(\d+)%", comment)
    footprint_match = re.search(r"([0-9]{4})", comment)

    if (
        voltage_match
        and value_match
        and type_match
        and tolerance_match
        and footprint_match
    ):
        voltage = float(voltage_match.group(1))
        value = f"{value_match.group(1)}{value_match.group(2)}"
        cap_type = type_match.group(1)
        tolerance = int(tolerance_match.group(1))
        footprint = footprint_match.group(1)
        key = (voltage, value, cap_type, tolerance, footprint)
        mpn = CAPACITOR_MAP[key]
        comment = f"{value} ±{tolerance}% {cap_type} {footprint} capacitor (MLCC)"
        data = {
            "mpn": mpn,
            "value": value,
            "footprint": footprint,
            "description": comment,
            "vendor": "DigiKey",
            "manufacturer": "Murata",
        }
        return data
    raise ValueError("Invalid capacitor comment")


def parse_bom_row(row, board_quantity):
    data = {
        "mpn": row["LCSC"],
        "designators": row["Designator"],
        "quantity": len(row["Designator"].split(",")),
        "description": row["Comment"],
    }

    order_quantity = data["quantity"] * board_quantity
    if "Resistor" in data["description"]:
        data.update(parse_resistor_comment(data["description"]))
        data.update(search_digikey_info(data["mpn"], order_quantity))
    elif "Capacitor" in data["description"]:
        data.update(parse_capacitor_comment(data["description"]))
        data.update(search_digikey_info(data["mpn"], order_quantity))
    elif "Do not populate" in data["description"]:
        data.update(
            {
                "mpn": "DNI",
                "dni": "DNI",
                "description": "Do not install",
            }
        )
    else:
        while data["mpn"] in MPN_MAP:
            data["mpn"] = MPN_MAP[data["mpn"]]
        data.update(search_digikey_info(data["mpn"], order_quantity))
    if "unit_price" in data:
        data["total_price"] = data["unit_price"] * order_quantity
    return data


def parse_bom(file, board_quantity):
    with open(file, "r") as f:
        reader = csv.DictReader(f)
        data = [parse_bom_row(row, board_quantity) for row in reader]
    return data


# parse bom path and board count as args:


def main():
    parser = argparse.ArgumentParser(description="Generate part library")
    parser.add_argument("--bom", help="Path to BOM file", required=True)
    parser.add_argument("--boards", type=int, help="Number of boards", default=1)
    parser.add_argument("--sierra", help="Path to Sierra BOM file", default=None)
    args = parser.parse_args()

    board_count = args.boards
    data = parse_bom(args.bom, board_count)
    data.sort(key=lambda x: float(x.get("total_price", 0.0)), reverse=True)

    SIMPLE_COLUMNS = {
        "MPN": "mpn",
        "Quantity": "quantity",
        "DNI": "dni",
        "Manufacturer": "manufacturer",
        "Vendor": "vendor",
        "Vendor Part Number": "vendor_part_number",
        "Value": "value",
        "Footprint": "footprint",
        "Description": "description",
        "Available": "available",
        "Unit Price": "unit_price",
        "Total Price": "total_price",
    }
    SIERRA_COLUMNS = {
        "Quantity per board": "quantity",
        "Manufacturer part number (MPN)": "mpn",
        "Reference designators": "designators",
        "DNI/DNP": "dni",
        "Vendor": "vendor",
        "Vendor part number": "vendor_part_number",
        "Value": "value",
        "Size/footprint": "footprint",
        "Part description/specs": "description",
        "Manufacturer": "manufacturer",
    }

    simple_data = []
    for row in data:
        simple_row = []
        for _, name in SIMPLE_COLUMNS.items():
            value = row.get(name, "")
            # Format currency values
            if name in ["unit_price", "total_price"] and value != "":
                value = f"${value:.2f}"
            simple_row.append(value)
        simple_data.append(simple_row)
    print(
        tabulate(
            simple_data,
            headers=list(SIMPLE_COLUMNS.keys()),
            tablefmt="rounded_outline",
        )
    )

    total_price = sum(row.get("total_price", 0.0) for row in data)
    total_price_per_board = total_price / board_count
    print(
        f"Total Price: ${total_price:.2f} (${total_price_per_board:.2f}/board * {board_count} boards)"
    )

    if args.sierra:
        sierra_table_data = []
        for row in data:
            sierra_row = []
            for _, name in SIERRA_COLUMNS.items():
                sierra_row.append(row.get(name, ""))
            sierra_table_data.append(sierra_row)
        # print(
        #     tabulate(
        #         sierra_table_data,
        #         headers=list(SIERRA_COLUMNS.keys()),
        #         tablefmt="rounded_outline",
        #     )
        # )

        # with open("sierra_bom.csv", "w", newline="") as csvfile:
        #     writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
        #     writer.writerow(list(SIERRA_COLUMNS.keys()))
        #     writer.writerows(sierra_table_data)

        # Create workbook and get active worksheet
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("BOM")
        else:
            ws.title = "BOM"

        # Write headers
        ws.append(list(SIERRA_COLUMNS.keys()))

        # Write data
        for row in sierra_table_data:
            # Convert first column to int, rest to strings
            formatted_row = [
                int(row[0]),  # quantity as number
                *[str(val) for val in row[1:]],  # rest as strings
            ]
            ws.append(formatted_row)

        # Adjust column widths automatically
        for idx, col in enumerate(ws.columns, start=1):  # Start counting from 1
            max_length = 0
            column = [cell for cell in col]
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

        # Save the file
        wb.save(args.sierra)


if __name__ == "__main__":
    main()
